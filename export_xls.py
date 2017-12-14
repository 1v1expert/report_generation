import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side
import get_point_list
import get_if
import import_xls

def load_file():
	#puth = 'consignors_dumps.json'
	puth = 'consignors_temp_' + str(get_if.get_time(True)) + 'json'
	#puth = 'consignors_dumps_2.json'
	with open(puth) as file:
		data = json.load(file)
		#print(data)
		chang_xls(data)
		file.close()

def chang_xls(data):
	#from openpyxl import Workbook

	fill = PatternFill(fill_type='solid',
                   start_color='c1c1c1',
                   end_color='c2c2c2')
	align_center=Alignment(horizontal='center',
                       vertical='bottom',
                       text_rotation=0,
                       wrap_text=False,
                       shrink_to_fit=True,
                       indent=0)
	border = Border(left=Side(border_style='thin',
                             color='FF000000'),
                   right=Side(border_style='thin',
                              color='FF000000'),
                   top=Side(border_style='thin',
                            color='FF000000'),
                   bottom=Side(border_style='thin',
                               color='FF000000'),
                   diagonal=Side(border_style='thin',
                                 color='FF000000'),
                   diagonal_direction=0,
                   outline=Side(border_style='thin',
                                color='FF000000'),
                   vertical=Side(border_style='thin',
                                 color='FF000000'),
                   horizontal=Side(border_style='thin',
                                  color='FF000000')
                  )
	font = Font(name='Times New Roman',
	                    size=10,
	                    bold=False,
	                    italic=True,
	                    vertAlign=None,
	                    underline='none',
	                    strike=False,
	                    color='FF000000')
	font2 = Font(name='Times New Roman',
	                    size=12,
	                    bold=True,
	                    italic=True,
	                    vertAlign=None,
	                    underline='none',
	                    strike=False,
	                    color='FF000000')
	wb = Workbook()
	ws= wb.active
	ws.title = 'consignors'
	#ws = wb.create_sheet(title = 'consignors')
	#font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
	#ws.append(['Consignors: '])
	#ws['A1'].fill = fill

	#ws.Font(italic = True)
	ws.append(['№ ордера', '№ точки в ДПД', '№ точки', '№ парсела', 'Время', 'Дата', 'Отправитель', 'Время выдачи', 'Дата выдачи', 'Время доставки', 'Дата доставки','В доставке(время)','В доставке(дата)', 'Ссылка в ИС'])
	for cellObj in ws['A1:N1']:
		for cell in cellObj:
			ws[cell.coordinate].font = font2
			ws[cell.coordinate].fill = fill
			ws[cell.coordinate].alignment = align_center
			ws[cell.coordinate].border = border

	i = 2
	langs = 1
	count = data['count']
	#for langs in data.keys():
	while langs<count+1:
		try:
			ws.append([data[str(langs)]['dpd_order'], data[str(langs)]['stationNums'], data[str(langs)]['pointCode'], data[str(langs)]['parcelNum:'], data[str(langs)]['Time'], data[str(langs)]['Date'], data[str(langs)]['consignor'], data[str(langs)]['issued_time'], data[str(langs)]['issued_date'], data[str(langs)]['delivered_time'], data[str(langs)]['delivered_date'], data[str(langs)]['delivering_time'], data[str(langs)]['delivering_date'], data[str(langs)]['ref']])
			for cellObj in ws['A' + str(i) + ':N' + str(i)]:
				for cell in cellObj:
					ws[cell.coordinate].border = border
					ws[cell.coordinate].alignment = align_center
					ws[cell.coordinate].font = font
			i += 1

		except:
			ws.append([data[str(langs)]['dpd_order'], data[str(langs)]['stationNums'], data[str(langs)]['pointCode'], data[str(langs)]['parcelNum:'], data[str(langs)]['Time'], data[str(langs)]['Date'], data[str(langs)]['consignor']])
			for cellObj in ws['A' + str(i) + ':N' + str(i)]:
				for cell in cellObj:
					ws[cell.coordinate].border = border
					ws[cell.coordinate].alignment = align_center
					ws[cell.coordinate].font = font
			i += 1


		#print(langs, '\t', data[langs]['stationNums'], '\t',data[langs]['pointCode'], '\t', data[langs]['parcelNum:'], '\t', data[langs]['Time'], '\t' , data[langs]['Date'], '\t', data[langs]['consignor'])
		langs += 1
#	for cellObj in ws['A' + str(i) + ':G' + str(i)]:
#		for cell in cellObj:
#			ws[cell.coordinate].border = border
#			ws[cell.coordinate].alignment = align_center
#			ws[cell.coordinate].font = font
#for cellObj in ws['A:G13']:
		#for cell in cellObj:
			#ws[cell.coordinate].border = border
			#ws[cell.coordinate].alignment = align_center
			#ws[cell.coordinate].font = font
			#ws[cell.coordinate].fill = fill
	"""col = ws.column_dimensions['A']
				col.font = Font(italic=True)
				row = ws.row_dimensions[1]
				row.font = Font(underline="single")"""
	
	#ws['A4'].alignment = align_center
	#ws['A4'].fill = fill
	#ws['A4'].width = 150
	wb.save('report.xlsx')
if __name__ == '__main__':
	get_if.rewritingg()
	get_if.get_parcel_pulse()
	#get_point_list.get_parcel_pulse()
	load_file()
	import_xls.load_data()

#'Поддержка т. 2134532'
